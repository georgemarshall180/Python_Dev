##
##  python summarizer.py -i /Users/smoon/Desktop/AAT_FrontPage.xlsx -o /Users/smoon/Desktop/Master.xlsx
##  Basically “-i” param is for your given input file, and “-o” param is your output file (master)
##  Here is the example command.
##  $ python summarizer.py -i /Users/smoon/Desktop/AAT_FrontPage.xlsx -o /Users/smoon/Desktop/Master.xlsx
##  need “openpyxl” python library on your python build path. download from https://pypi.python.org/pypi/openpyxl
##   


import sys, getopt
import openpyxl
import re
from openpyxl.worksheet import Worksheet


def params(argv):
    input_file = ''
    output_file = ''
    summary_require = False;
    try:
        opts, args = getopt.getopt(argv, "hi:o:s", ["ifile=", "ofile=", "summary_require"])
    except getopt.GetoptError:
        print('test.py -i <inputfile> -o <outputfile>')
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print('test.py -i <inputfile> -o <outputfile>')
            sys.exit()
        elif opt in ("-i", "--ifile"):
            input_file = arg
        elif opt in ("-o", "--ofile"):
            output_file = arg
        elif opt in ("-s", "--summary_require"):
            summary_require = True

    print("Input file is \"", input_file)
    print("Output file is \"", output_file)

    # No summary update(Still working on it, not yet implemented)
    summary_require = False;

    return {summary_require, input_file, output_file}


def copy_sheets(given_summary_require, given_input_file, given_output_file):
    input_excel = openpyxl.load_workbook(given_input_file)
    output_excel = openpyxl.load_workbook(given_output_file)

    print("Output Excel file's sheet list:")
    output_file_sheets = output_excel.get_sheet_names()
    print(output_file_sheets)
    if output_file_sheets[0] == "Sheet1":
        output_excel.remove_sheet(output_excel.get_sheet_by_name('Sheet1'))
    elif output_file_sheets[0] == "Sheet":
        output_excel.remove_sheet(output_excel.get_sheet_by_name('Sheet'))

    output_file_sheets = output_excel.get_sheet_names()
    print(output_file_sheets)

    # input_file_sheets = input_excel.get_sheet_names()

    # for input_sheet in input_file_sheets:
    for input_sheet in input_excel:
        # input_file_sheets = input_excel.get_sheet_names()

        type(input_sheet)
        sheet_name = input_sheet.title
        sheet_exists_in_output = False

        for index in range(len(output_file_sheets)):
            if output_file_sheets[index] == sheet_name:
                sheet_exists_in_output = True
                print("Sheet ", sheet_name, " already EXIST!")
                output_excel.remove_sheet(output_excel.get_sheet_by_name(sheet_name))

        # Needs to copy all contencts!
        if sheet_exists_in_output == False:
            print(sheet_name, " is created and copied!")
            output_sheet = output_excel.create_sheet(title=sheet_name)

            # Trial...
            for row_of_cell_objs in input_sheet.iter_rows():
                for cell_obj in row_of_cell_objs:
                    output_sheet[cell_obj.coordinate] = cell_obj.internal_value
                    output_sheet[cell_obj.coordinate].style = input_sheet[cell_obj.coordinate].style
                    # output_sheet[cell_obj.coordinate] = output_sheet.cell(row=cell_obj.row, col_idx=cell_obj.col_idx, value=cell_obj.value)
                    if cell_obj.has_style:
                        output_sheet[cell_obj.coordinate].font = cell_obj.font
                        output_sheet[cell_obj.coordinate].border = cell_obj.border
                        output_sheet[cell_obj.coordinate].fill = cell_obj.fill
                        output_sheet[cell_obj.coordinate].number_format = cell_obj.number_format
                        output_sheet[cell_obj.coordinate].protection = cell_obj.protection
                        output_sheet[cell_obj.coordinate].alignment = cell_obj.alignment

            if given_summary_require:
                # Get last index from Summary sheet to add the application info.
                summary_sheet = output_excel.get_sheet_by_name('Summary')
                # It search only from 1 to 100 lines, so if the app size are bigger than 100, should increase!
                # Needs to summarize the sheet on Summary sheet!
                next_available_index_on_summary = find_next_line_on_summary_sheet(summary_sheet)
                print("You requested to update on SUMMARY sheet")
                insert_result_in_summary_sheet(output_excel, input_sheet, next_available_index_on_summary)

    # To Save what it has been done.
    output_excel.save(given_output_file)


def find_next_line_on_summary_sheet(summary_sheet):
    # It search only from 1 to 100 lines, so if the app size are bigger than 100, should increase!
    next_available_index_on_summary = 0
    # for row in summary_sheet.iter_rows():

    # Good approach
    # for first_cells in summary_sheet['A1':'J2']:
    for first_cells in summary_sheet.iter_rows():
        for first_cell in first_cells:
            if first_cell.internal_value == 'L1 average':
                next_available_index_on_summary = first_cell.row
                print("FOUND!!!")
                return next_available_index_on_summary


# Comment block starts..
# ...
def insert_result_in_summary_sheet(output_excel, input_sheet, next_available_index_on_summary):
    old_summary_sheet = output_excel.get_sheet_by_name("Summary")
    max_row = old_summary_sheet.get_highest_row()
    max_col = old_summary_sheet.get_highest_column()
    old_summary_sheet.title = "Summary_OLD"
    output_excel.create_sheet(0, 'Summary')

    new_summary_sheet = output_excel.get_sheet_by_name('Summary')

    # Do the header.
    for col_num in range(0, max_col):
        new_summary_sheet.cell(row=0, column=col_num).value = old_summary_sheet.cell(row=0, column=col_num).value

    is_new_col_added = False
    for row_num in range(1, max_row):
        if row_num == next_available_index_on_summary:
            is_new_col_added = True
            new_row_num = row_num - 1

            new_summary_sheet.cell(row=row_num, column=0).value = row_num - 1
            new_summary_sheet.cell(row=row_num, column=1).value = input_sheet.title
            new_summary_sheet.cell(row=row_num, column=2).value = 1
            new_summary_sheet.cell(row=row_num, column=3).value = None
            new_summary_sheet.cell(row=row_num, column=4).set_value_explicit(value="=" + input_sheet.title + "D32",
                                                                             data_type='f')
            new_summary_sheet.cell(row=row_num, column=5).value = new_summary_sheet.cell(row=row_num, column=4).value
            new_summary_sheet.cell(row=row_num, column=6).value = '20%'
            new_summary_sheet.cell(row=row_num, column=7).set_value_explicit(
                value="=IF(E" + row_num + "=0,G" + row_num + ",E" + row_num + "*(1+G" + row_num + "))", data_type='f')
            new_summary_sheet.cell(row=row_num, column=8).set_value_explicit(
                value="=F" + row_num + "-E" + row_num + ")/E" + row_num, data_type='f')
            new_summary_sheet.cell(row=row_num, column=9).value = None
            new_summary_sheet.cell(row=row_num, column=10).set_value_explicit(value="=" + input_sheet.title + "E8",
                                                                              data_type='f')

        else:
            if is_new_col_added:  # L1 Average & GOAL
                if row_num == next_available_index_on_summary + 1:  # L1 Avg
                    print(" what")
                else:  # GOAL
                    print(" something...")
            else:  # Normal case...
                for col_num in range(0, max_col):
                    if col_num == 4 or col_num == 7 or col_num == 8 or col_num == 10:
                        new_summary_sheet.cell(row=row_num, column=col_num).set_value_explicit(
                            value=old_summary_sheet.cell(row=row_num, column=col_num).value, data_type='f')
                    else:
                        new_summary_sheet.cell(row=row_num, column=col_num).value = old_summary_sheet.cell(row=row_num,
                                                                                                           column=col_num).value


# '''
# Comment block end

if __name__ == "__main__":
    summary_r, outputf, inputf = params(sys.argv[1:])
    # outputf, inputf, summary_r= params(sys.argv[1:])
    #	Worksheet.insert_rows = insert_rows
    copy_sheets(summary_r, inputf, outputf)
